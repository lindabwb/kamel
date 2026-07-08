"""
Extraction de certificats qualite PCB depuis des PDF.

Usage:
    python pcb.py "C:\\chemin\\vers\\dossier_pdf" --output rapport_pcb.xlsx
    python pcb.py "C:\\chemin\\vers\\dossier_pdf" --output rapport_pcb.csv

Dependances:
    pip install pdfplumber openpyxl
"""

from __future__ import annotations

import argparse
import csv
import os
import re
import sys
import threading
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Optional

import pdfplumber


COVER_COLUMNS = [
    "FileName",
    "Champ",
    "Valeur nom fichier",
    "Valeur page de garde",
    "Comparaison",
]

STANDARD_COLUMNS = [
    "FileName",
    "Norme",
    "Source",
]

INSPECTION_COLUMNS = [
    "FileName",
    "TestName",
    "SPEC",
    "RESULTS",
    "Conformite",
    "Commentaire",
]

FIELD_PATTERNS = {
    "P.O.NO": [
        r"\bP\.?\s*O\.?\s*NO\.?\s*[:：]?\s*([A-Z0-9._/\-]+)",
        r"\bPO\s*NO\.?\s*[:：]?\s*([A-Z0-9._/\-]+)",
    ],
    "PART NO": [
        r"\bPART\s*(?:NO|NUMBER)\.?\s*[:：]?\s*([A-Z0-9._/\-]+)",
    ],
    "QUANTITY": [
        r"\bQUANTITY\s*[:：]?\s*([0-9][0-9,.\s]*)(?:\s*PCS)?",
    ],
    "DATA CODE": [
        r"\bDATA\s*CODE\s*[:：]?\s*([A-Z0-9._/\-]+)",
        r"\bDATE\s*CODE\s*[:：]?\s*([A-Z0-9._/\-]+)",
    ],
}

FIELD_LABEL_RE = re.compile(
    r"\b(P\.?\s*O\.?\s*NO\.?|PO\s*NO\.?|PART\s*(?:NO|NUMBER)\.?|QUANTITY|DATA\s*CODE|DATE\s*CODE|CUSTOMER|COMPONENT|APPROVED\s*DATE|SERIAL\s*NO)\b",
    flags=re.IGNORECASE,
)

KNOWN_STANDARD_PATTERNS = [
    r"\bUL\s*94\s*Flame\s*Class\s*94?V-?0\b",
    r"\bUL\s*94\s*V-?0\b",
    r"\b94V-?0\b",
    r"\bRoHS\b(?:\s+Directive)?",
]

SKIP_PAGE_MARKERS = [
    "Product iQ",
    "Wiring, Printed - Component",
]


@dataclass
class CoverData:
    po_no: str = "NA"
    part_no: str = "NA"
    quantity: str = "NA"
    data_code: str = "NA"
    standards: str = "NA"


@dataclass
class PdfReport:
    cover_rows: list[dict[str, str]]
    standard_rows: list[dict[str, str]]
    inspection_rows: list[dict[str, str]]
    warnings: list[str]


def value_or_na(value: str) -> str:
    value = clean_text(value)
    return value if value else "NA"


def clean_text(value: object) -> str:
    """Normalize extracted PDF cell text."""
    if value is None:
        return ""
    text = str(value).replace("\n", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def normalize_word(value: str) -> str:
    return re.sub(r"[^A-Z0-9]+", "", value.upper())


def group_words_by_line(words: list[dict[str, object]], tolerance: float = 3.0) -> list[list[dict[str, object]]]:
    lines: list[list[dict[str, object]]] = []
    for word in sorted(words, key=lambda item: (float(item["top"]), float(item["x0"]))):
        top = float(word["top"])
        for line in lines:
            if abs(float(line[0]["top"]) - top) <= tolerance:
                line.append(word)
                break
        else:
            lines.append([word])
    for line in lines:
        line.sort(key=lambda item: float(item["x0"]))
    return lines


def find_label_index(normalized_words: list[str], label_words: list[str]) -> Optional[int]:
    for index in range(0, len(normalized_words) - len(label_words) + 1):
        if normalized_words[index : index + len(label_words)] == label_words:
            return index
    return None


def pick_cover_value(candidate_words: list[str], field_name: str) -> str:
    cleaned = [
        clean_text(word)
        for word in candidate_words
        if normalize_word(word) != "" and clean_text(word) not in {":", "："}
    ]
    if field_name == "QUANTITY":
        number = next((re.search(r"\d+", word).group(0) for word in cleaned if re.search(r"\d+", word)), "")
        unit = next((word for word in cleaned if normalize_word(word) in {"PCS", "PIECES"}), "")
        return clean_text(f"{number} {unit}") if number else "NA"

    if field_name in {"P.O.NO", "QUANTITY", "DATA CODE"}:
        for word in cleaned:
            match = re.search(r"\d+", word)
            if match:
                return clean_quantity(match.group(0)) if field_name == "QUANTITY" else match.group(0)
    if field_name == "PART NO":
        for word in cleaned:
            if re.search(r"[A-Z]", word, flags=re.IGNORECASE) and re.search(r"\d", word):
                return word
    return cleaned[0] if cleaned else "NA"


def pick_cover_value_by_coordinates(
    words: list[dict[str, object]],
    label_words: list[dict[str, object]],
    field_name: str,
) -> str:
    label_x1 = max(float(word["x1"]) for word in label_words)
    label_top = min(float(word["top"]) for word in label_words)
    label_bottom = max(float(word["bottom"]) for word in label_words)
    label_center = (label_top + label_bottom) / 2

    candidates: list[tuple[float, float, str]] = []
    for word in words:
        text = str(word["text"])
        normalized = normalize_word(text)
        if normalized == "" or text in {":", "："}:
            continue
        x0 = float(word["x0"])
        x1 = float(word["x1"])
        center = (float(word["top"]) + float(word["bottom"])) / 2
        if x0 <= label_x1 + 40:
            continue
        if x0 < 250:
            continue
        if abs(center - label_center) > 16:
            continue
        candidates.append((abs(center - label_center), x0, text))

    candidates.sort()
    return pick_cover_value([text for _, _, text in candidates], field_name)


def extract_cover_fields_by_position(pdf: pdfplumber.PDF) -> dict[str, str]:
    labels = {
        "P.O.NO": [["PONO"], ["PO", "NO"]],
        "PART NO": [["PARTNO"], ["PART", "NO"]],
        "QUANTITY": [["QUANTITY"]],
        "DATA CODE": [["DATACODE"], ["DATA", "CODE"]],
    }
    fields = {field: "NA" for field in labels}

    for page in pdf.pages:
        text = page.extract_text(x_tolerance=1, y_tolerance=3) or ""
        if not re.search(r"APPROVED\s+CERTIFICATION|CERTIFICATION\s+NO|P\.?\s*O\.?\s*NO", text, flags=re.IGNORECASE):
            continue

        words = page.extract_words(use_text_flow=False, keep_blank_chars=False)
        all_words = words
        for line in group_words_by_line(words):
            line_words = [str(word["text"]) for word in line]
            normalized_words = [normalize_word(word) for word in line_words]
            for field_name, label_variants in labels.items():
                if fields[field_name] != "NA":
                    continue
                found_label: Optional[tuple[int, list[str]]] = None
                for label_words in label_variants:
                    label_index = find_label_index(normalized_words, label_words)
                    if label_index is not None:
                        found_label = (label_index, label_words)
                        break
                if found_label is None:
                    continue
                label_index, label_words = found_label
                label_word_objects = line[label_index : label_index + len(label_words)]
                fields[field_name] = pick_cover_value_by_coordinates(
                    all_words, label_word_objects, field_name
                )

        if all(value != "NA" for value in fields.values()):
            break

    return fields


def clean_quantity(value: str) -> str:
    value = clean_text(value).upper().replace(",", "")
    value = re.sub(r"\bPCS\b", "", value).strip()
    return re.sub(r"\s+", "", value) or "NA"


def normalize_for_compare(value: str, field_name: str) -> str:
    value = clean_text(value).upper()
    if field_name == "QUANTITY":
        return "".join(re.findall(r"\d+", value))
    return re.sub(r"[^A-Z0-9]", "", value)


def to_float(value: str) -> Optional[float]:
    match = re.search(r"-?\d+(?:[.,]\d+)?", value)
    if not match:
        return None
    return float(match.group(0).replace(",", "."))


def extract_numbers(value: str) -> list[float]:
    return [float(number.replace(",", ".")) for number in re.findall(r"(?<!\d)-?\d+(?:[.,]\d+)?", value)]


def parse_filename_values_from_name(filename: str) -> dict[str, str]:
    """
    Expected file name pattern:
    PARTNO-PONO-QUANTITY-DATACODE, accepting hyphen/en-dash/em-dash separators.

    Extra suffixes after the fourth token are ignored, e.g. supplier name.
    """
    stem = Path(filename).stem
    if "—" in stem or "–" in stem:
        parts = [p.strip() for p in re.split(r"\s*[–—]\s*", stem) if p.strip()]
    else:
        parts = [p.strip() for p in re.split(r"\s*-\s*", stem) if p.strip()]
    values = {"PART NO": "NA", "P.O.NO": "NA", "QUANTITY": "NA", "DATA CODE": "NA"}
    if len(parts) >= 4:
        values["PART NO"] = parts[0]
        values["P.O.NO"] = parts[1]
        values["QUANTITY"] = clean_quantity(parts[2])
        values["DATA CODE"] = parts[3]
    return values


def parse_filename_values(pdf_path: Path) -> dict[str, str]:
    return parse_filename_values_from_name(pdf_path.name)


def regex_first(patterns: Iterable[str], text: str) -> str:
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE | re.MULTILINE)
        if match:
            return clean_text(match.group(1))
    return "NA"


def extract_field_between_labels(text: str, label: str, next_labels: list[str], field_name: str) -> str:
    label_pattern = re.escape(label).replace(r"\ ", r"\s+")
    next_pattern = "|".join(re.escape(item).replace(r"\ ", r"\s+") for item in next_labels)
    pattern = rf"\b{label_pattern}\b\s*[:：]?(?P<value>.*?)(?=\b(?:{next_pattern})\b\s*[:：]?|$)"
    match = re.search(pattern, text, flags=re.IGNORECASE | re.DOTALL)
    if not match:
        return "NA"

    block = clean_text(match.group("value"))
    block = re.sub(r"[_—\-]{2,}", " ", block)
    block = re.sub(r"\bPCS\b", " ", block, flags=re.IGNORECASE)
    tokens = re.findall(r"[A-Z0-9]+(?:[./_-][A-Z0-9]+)*", block, flags=re.IGNORECASE)
    if not tokens:
        return "NA"

    if field_name in {"P.O.NO", "DATA CODE"}:
        for token in tokens:
            if re.fullmatch(r"\d+", token):
                return token
    if field_name == "QUANTITY":
        for token in tokens:
            if re.fullmatch(r"\d+", token):
                return clean_quantity(token)
    if field_name == "PART NO":
        for token in tokens:
            if re.search(r"[A-Z]", token, flags=re.IGNORECASE) and not re.fullmatch(r"\d{4}/\d{1,2}/\d{1,2}", token):
                return token

    return tokens[0]


def extract_fields_from_cover_block(all_text: str) -> dict[str, str]:
    # The first certification page has visual lines, so values can sit far from labels.
    match = re.search(
        r"MASS\s+PRODUCTION\s+APPROVED\s+CERTIFICATION(?P<body>.*?)(?:CUSTOMER\s+RECOGNITION|Contents|CERTIFICATE\s+OF\s+COMPLIANCE)",
        all_text,
        flags=re.IGNORECASE | re.DOTALL,
    )
    cover_text = match.group("body") if match else all_text
    return {
        "P.O.NO": extract_field_between_labels(
            cover_text, "P.O.NO", ["COMPONENT", "PART NO", "APPROVED DATE", "SERIAL NO", "QUANTITY", "DATA CODE"], "P.O.NO"
        ),
        "PART NO": extract_field_between_labels(
            cover_text, "PART NO", ["APPROVED DATE", "SERIAL NO", "QUANTITY", "DATA CODE"], "PART NO"
        ),
        "QUANTITY": extract_field_between_labels(
            cover_text, "QUANTITY", ["DATA CODE", "CUSTOMER RECOGNITION", "Contents"], "QUANTITY"
        ),
        "DATA CODE": extract_field_between_labels(
            cover_text, "DATA CODE", ["CUSTOMER RECOGNITION", "Contents", "CERTIFICATE OF COMPLIANCE"], "DATA CODE"
        ),
    }


def canonical_field_name(label: str) -> Optional[str]:
    label_key = re.sub(r"[^A-Z]", "", label.upper())
    if label_key in {"PONO", "PONO"}:
        return "P.O.NO"
    if label_key in {"PARTNO", "PARTNUMBER"}:
        return "PART NO"
    if label_key == "QUANTITY":
        return "QUANTITY"
    if label_key in {"DATACODE", "DATECODE"}:
        return "DATA CODE"
    return None


def extract_value_after_label(text: str, label: str) -> str:
    """
    Extract a field value even when the PDF puts the value on the next line.

    The first page often looks like:
        PART NO :

            CAD00000085C
    Simple regexes can accidentally jump to the next unrelated label; this
    routine stops as soon as another known label is reached.
    """
    label_pattern = re.escape(label).replace(r"\ ", r"\s*")
    match = re.search(label_pattern + r"\s*[:：]?", text, flags=re.IGNORECASE)
    if not match:
        return "NA"

    tail = text[match.end() :]
    values: list[str] = []
    for raw_line in tail.splitlines():
        line = clean_text(raw_line)
        if not line:
            continue
        if FIELD_LABEL_RE.search(line):
            break
        values.append(line)
        break

    if not values:
        return "NA"

    value = values[0]
    value = re.sub(r"\bPCS\b", "", value, flags=re.IGNORECASE).strip()
    return value or "NA"


def extract_fields_from_text(all_text: str) -> dict[str, str]:
    field_labels = {
        "P.O.NO": ["P.O.NO", "P.O. NO", "PO NO"],
        "PART NO": ["PART NO", "PART NUMBER"],
        "QUANTITY": ["QUANTITY"],
        "DATA CODE": ["DATA CODE", "DATE CODE"],
    }
    fields: dict[str, str] = extract_fields_from_cover_block(all_text)
    for field_name, labels in field_labels.items():
        if fields.get(field_name, "NA") != "NA":
            continue
        value = "NA"
        for label in labels:
            value = extract_value_after_label(all_text, label)
            if value != "NA":
                break
        if value == "NA":
            value = regex_first(FIELD_PATTERNS[field_name], all_text)
        fields[field_name] = value
    fields["QUANTITY"] = clean_quantity(fields["QUANTITY"])
    return fields


def extract_cover_data(all_text: str) -> CoverData:
    fields = extract_fields_from_text(all_text)
    standards = extract_standards(all_text)
    return CoverData(
        po_no=fields["P.O.NO"],
        part_no=fields["PART NO"],
        quantity=fields["QUANTITY"],
        data_code=fields["DATA CODE"],
        standards=standards,
    )


def extract_standards(all_text: str) -> str:
    certificate_text = all_text
    match = re.search(
        r"CERTIFICATE\s+OF\s+COMPLIANCE(?P<body>.*?)(?:Product\s+iQ|INSPECTION\s+REPORT|$)",
        all_text,
        flags=re.IGNORECASE | re.DOTALL,
    )
    if match:
        certificate_text = match.group("body")

    found: list[str] = []
    for pattern in KNOWN_STANDARD_PATTERNS:
        for match in re.finditer(pattern, certificate_text, flags=re.IGNORECASE):
            standard = clean_text(match.group(0))
            if re.search(r"UL\s*94|94V-?0", standard, flags=re.IGNORECASE):
                standard = "UL 94V-0"
            if standard.lower().startswith("rohs"):
                standard = "RoHS Directive"
            if standard not in found:
                found.append(standard)

    if not found:
        for pattern in KNOWN_STANDARD_PATTERNS:
            for match in re.finditer(pattern, all_text, flags=re.IGNORECASE):
                standard = clean_text(match.group(0))
                if re.search(r"UL\s*94|94V-?0", standard, flags=re.IGNORECASE):
                    standard = "UL 94V-0"
                if standard.lower().startswith("rohs"):
                    standard = "RoHS Directive"
                if standard not in found:
                    found.append(standard)
    return "; ".join(found) if found else "NA"


def split_standards(standards: str) -> list[str]:
    if standards == "NA":
        return ["NA"]
    return [item.strip() for item in standards.split(";") if item.strip()]


def cover_to_dict(cover: CoverData) -> dict[str, str]:
    return {
        "P.O.NO": cover.po_no,
        "PART NO": cover.part_no,
        "QUANTITY": cover.quantity,
        "DATA CODE": cover.data_code,
    }


def build_cover_comparison_rows(
    pdf_path: Path,
    cover: CoverData,
    display_name: Optional[str] = None,
) -> tuple[list[dict[str, str]], list[str]]:
    file_name = display_name or pdf_path.name
    expected = parse_filename_values_from_name(file_name)
    extracted = cover_to_dict(cover)
    rows: list[dict[str, str]] = []
    warnings: list[str] = []

    for field_name in ["PART NO", "P.O.NO", "QUANTITY", "DATA CODE"]:
        expected_value = expected.get(field_name, "NA")
        extracted_value = extracted.get(field_name, "NA")
        if expected_value == "NA" or extracted_value == "NA":
            status = "A VERIFIER"
        elif normalize_for_compare(extracted_value, field_name) == normalize_for_compare(expected_value, field_name):
            status = "OK"
        else:
            status = "DIFFERENT"
            warnings.append(
                f"{file_name}: {field_name} PDF='{extracted_value}' fichier='{expected_value}'"
            )

        rows.append(
            {
                "FileName": file_name,
                "Champ": field_name,
                "Valeur nom fichier": expected_value,
                "Valeur page de garde": extracted_value,
                "Comparaison": status,
            }
        )

    return rows, warnings


def build_standard_rows(pdf_path: Path, standards: str, display_name: Optional[str] = None) -> list[dict[str, str]]:
    file_name = display_name or pdf_path.name
    return [
        {
            "FileName": file_name,
            "Norme": standard,
            "Source": "Certificate of Compliance",
        }
        for standard in split_standards(standards)
    ]


def evaluate_conformity(spec: str, result: str) -> tuple[str, str]:
    spec_clean = clean_text(spec)
    result_clean = clean_text(result)
    numeric_spec = re.sub(r"\b[A-Z]+\d+\s*:", "", spec_clean, flags=re.IGNORECASE)
    spec_upper = spec_clean.upper()
    result_upper = result_clean.upper()

    if spec_upper == "NA" and result_upper == "NA":
        return "CONFORME", "Non applicable dans la specification et le resultat"

    if not spec_clean or not result_clean or spec_upper == "NA" or result_upper == "NA":
        return "A VERIFIER", "Specification ou resultat manquant/NA"

    negative_words = ["FAIL", "FAILED", "NG", "NOK", "NOT OK", "REJECT", "REJECTED"]
    if any(word in result_upper for word in negative_words):
        return "NON CONFORME", "Le resultat indique un echec"

    if normalize_for_compare(spec_clean, "") == normalize_for_compare(result_clean, ""):
        return "CONFORME", "Resultat identique a la specification"

    result_numbers = extract_numbers(result_clean)
    spec_numbers = extract_numbers(numeric_spec)

    if result_numbers and spec_numbers and re.search(r"\bMIN\.?\b|>=|≧|≥", spec_upper):
        limit = spec_numbers[0]
        failing = [value for value in result_numbers if value < limit]
        return (
            ("CONFORME", f"Toutes les mesures >= {limit:g}")
            if not failing
            else ("NON CONFORME", f"Mesure(s) < {limit:g}: {', '.join(f'{value:g}' for value in failing)}")
        )

    if result_numbers and spec_numbers and re.search(r"\bMAX\.?\b|<=|≦|≤", spec_upper):
        limit = spec_numbers[0]
        failing = [value for value in result_numbers if value > limit]
        return (
            ("CONFORME", f"Toutes les mesures <= {limit:g}")
            if not failing
            else ("NON CONFORME", f"Mesure(s) > {limit:g}: {', '.join(f'{value:g}' for value in failing)}")
        )

    if result_numbers and ("~" in numeric_spec or "～" in numeric_spec) and len(spec_numbers) >= 2:
        low, high = spec_numbers[0], spec_numbers[1]
        failing = [value for value in result_numbers if not low <= value <= high]
        return (
            ("CONFORME", f"Toutes les mesures dans la plage {low:g}-{high:g}")
            if not failing
            else ("NON CONFORME", f"Mesure(s) hors plage {low:g}-{high:g}: {', '.join(f'{value:g}' for value in failing)}")
        )

    if result_numbers and re.search(r"\d\s*-\s*\d", numeric_spec) and len(spec_numbers) >= 2:
        low, high = sorted((spec_numbers[0], spec_numbers[1]))
        failing = [value for value in result_numbers if not low <= value <= high]
        return (
            ("CONFORME", f"Toutes les mesures dans la plage {low:g}-{high:g}")
            if not failing
            else ("NON CONFORME", f"Mesure(s) hors plage {low:g}-{high:g}: {', '.join(f'{value:g}' for value in failing)}")
        )

    if result_numbers and re.search(r"\+/-|±", numeric_spec) and len(spec_numbers) >= 2:
        target, tolerance = spec_numbers[0], abs(spec_numbers[1])
        low, high = target - tolerance, target + tolerance
        failing = [value for value in result_numbers if not low <= value <= high]
        return (
            ("CONFORME", f"Toutes les mesures dans la tolerance {low:g}-{high:g}")
            if not failing
            else ("NON CONFORME", f"Mesure(s) hors tolerance {low:g}-{high:g}: {', '.join(f'{value:g}' for value in failing)}")
        )

    positive_words = ["OK", "PASS", "PASSED", "ACCEPT", "ACCEPTED", "CONFORM"]
    if any(word in result_upper for word in positive_words):
        return "CONFORME", "Le resultat indique OK/PASS"

    return "A VERIFIER", "Regle de specification non reconnue automatiquement"


def is_contents_page(text: str) -> bool:
    compact = re.sub(r"\s+", " ", text).strip().lower()
    return compact.startswith("contents") or (
        "contents" in compact
        and "certificate of compliance" in compact
        and "inspection report" in compact
        and len(compact) < 1200
    )


def should_skip_page(text: str) -> bool:
    if is_contents_page(text):
        return True
    return any(marker.lower() in text.lower() for marker in SKIP_PAGE_MARKERS)


def normalize_header(value: object) -> str:
    text = clean_text(value).upper()
    text = re.sub(r"[^A-Z0-9]+", "", text)
    if text in {"SPECIFICATION", "SPECIFICATIONS", "DRWDIMENSION", "DIMENSION", "STD", "STANDARD"}:
        return "SPEC"
    if text in {"RESULT", "RESULTS", "MEASUREDRESULTS", "THICKNESS"}:
        return "RESULTS"
    if text in {"REMARK", "REMARKS", "PTHNPTH"}:
        return "REMARKS"
    if text in {"NO", "NO."}:
        return "ITEM"
    return text


def find_header_row(table: list[list[object]]) -> Optional[tuple[int, Optional[int], int, list[int], list[int]]]:
    for row_index, row in enumerate(table):
        headers = [normalize_header(cell) for cell in row]
        description_index = next((i for i, h in enumerate(headers) if h == "DESCRIPTION"), None)
        spec_index = next((i for i, h in enumerate(headers) if h == "SPEC"), None)
        results_index = next((i for i, h in enumerate(headers) if h == "RESULTS"), None)
        if spec_index is not None and results_index is not None:
            spec_indexes = list(range(spec_index, results_index))
            results_indexes = [results_index]
            for i in range(results_index + 1, len(headers)):
                if headers[i] in {"REMARKS", "PTHNPTH", "TYPE", "VENDOR", "LOCATION"}:
                    break
                if headers[i] in {"", "RESULTS"}:
                    results_indexes.append(i)
            return row_index, description_index, spec_index, spec_indexes, results_indexes

        remarks_index = next((i for i, h in enumerate(headers) if h == "REMARKS"), None)
        if spec_index is not None and remarks_index is not None and remarks_index > spec_index:
            spec_indexes = [spec_index]
            results_indexes = list(range(spec_index + 1, remarks_index + 1))
            return row_index, description_index, spec_index, spec_indexes, results_indexes
    return None


def row_has_data(spec: str, results: str) -> bool:
    if not spec and not results:
        return False
    noise = {"SPEC", "RESULTS", "REMARKS"}
    return f"{spec} {results}".upper() not in noise


def is_test_name_noise(value: str) -> bool:
    value = clean_text(value)
    upper = value.upper()
    compact = normalize_header(value)
    if not value:
        return True
    if compact in {"ITEM", "DESCRIPTION", "TESTMETHOD", "NOOFSAMPLE", "SAMPLE", "SPEC", "RESULTS", "REMARKS"}:
        return True
    if re.fullmatch(r"\d+", upper):
        return True
    if re.search(r"\b(IPC|ANSI|J-STD|BELLCORE)\b", upper):
        return True
    if re.fullmatch(r"(NA|ALL|\d+\s*(ARRAY|PNL|PANEL)|\d+\s*PCS?)", upper):
        return True
    return False


def compose_test_name(raw_row: list[object], spec_col: int, current_parent: str) -> tuple[str, str]:
    cells_before_spec = [clean_text(cell) for cell in raw_row[:spec_col]]
    first_cell = cells_before_spec[0] if cells_before_spec else ""
    item_is_present = bool(first_cell and re.fullmatch(r"\d+", first_cell))
    parts = [cell for cell in cells_before_spec if not is_test_name_noise(cell)]

    if item_is_present and parts:
        current_parent = parts[0]
        return " - ".join(parts), current_parent

    if parts:
        if current_parent != "NA" and parts[0].upper() != current_parent.upper():
            return f"{current_parent} - {' - '.join(parts)}", current_parent
        return " - ".join(parts), current_parent

    if item_is_present:
        return f"ITEM {first_cell}", current_parent

    return current_parent, current_parent


def extract_spec_results_from_table(table: list[list[object]]) -> list[tuple[str, str, str]]:
    header = find_header_row(table)
    if header is None:
        return []

    header_row, description_col, spec_col, spec_cols, results_cols = header
    rows: list[tuple[str, str, str]] = []
    current_test = "NA"
    current_parent = "NA"
    previous_spec = ""

    for raw_row in table[header_row + 1 :]:
        if not raw_row:
            continue
        current_test, current_parent = compose_test_name(raw_row, spec_col, current_parent)
        if current_test == "NA" and description_col is not None and description_col < len(raw_row):
            description = clean_text(raw_row[description_col])
            if description and normalize_header(description) != "DESCRIPTION":
                current_test = description

        spec_values = [
            clean_text(raw_row[index])
            for index in spec_cols
            if index < len(raw_row) and clean_text(raw_row[index])
        ]
        spec = " ".join(spec_values)
        result_values = [
            clean_text(raw_row[index])
            for index in results_cols
            if index < len(raw_row) and clean_text(raw_row[index])
        ]
        results = " | ".join(result_values)
        if not spec and previous_spec and results:
            spec = previous_spec
        if spec:
            previous_spec = spec
        if row_has_data(spec, results):
            rows.append((current_test, spec or "NA", results or "NA"))

    return rows


def table_settings() -> list[dict[str, object]]:
    """Try line-based tables first, then text-position tables for borderless PDFs."""
    return [
        {
            "vertical_strategy": "lines",
            "horizontal_strategy": "lines",
            "snap_tolerance": 3,
            "join_tolerance": 3,
            "intersection_tolerance": 5,
        },
        {
            "vertical_strategy": "text",
            "horizontal_strategy": "text",
            "snap_tolerance": 3,
            "join_tolerance": 3,
            "intersection_tolerance": 5,
            "min_words_vertical": 2,
            "min_words_horizontal": 1,
        },
    ]


def extract_pdf_report(pdf_path: Path, display_name: Optional[str] = None) -> PdfReport:
    file_name = display_name or pdf_path.name
    inspection_rows: list[dict[str, str]] = []
    warnings: list[str] = []
    cover_rows: list[dict[str, str]] = []
    standard_rows: list[dict[str, str]] = []

    with pdfplumber.open(pdf_path) as pdf:
        page_texts = [(page.extract_text(x_tolerance=1, y_tolerance=3) or "") for page in pdf.pages]
        all_text = "\n".join(page_texts)
        cover = extract_cover_data(all_text)
        positioned_fields = extract_cover_fields_by_position(pdf)
        cover = CoverData(
            po_no=value_or_na(positioned_fields.get("P.O.NO", "NA")) if positioned_fields.get("P.O.NO", "NA") != "NA" else cover.po_no,
            part_no=value_or_na(positioned_fields.get("PART NO", "NA")) if positioned_fields.get("PART NO", "NA") != "NA" else cover.part_no,
            quantity=value_or_na(positioned_fields.get("QUANTITY", "NA")) if positioned_fields.get("QUANTITY", "NA") != "NA" else cover.quantity,
            data_code=value_or_na(positioned_fields.get("DATA CODE", "NA")) if positioned_fields.get("DATA CODE", "NA") != "NA" else cover.data_code,
            standards=cover.standards,
        )
        cover_rows, cover_warnings = build_cover_comparison_rows(pdf_path, cover, file_name)
        standard_rows = build_standard_rows(pdf_path, cover.standards, file_name)
        warnings.extend(cover_warnings)

        for page, text in zip(pdf.pages, page_texts):
            if should_skip_page(text):
                continue

            for settings in table_settings():
                page_rows: list[dict[str, str]] = []
                try:
                    tables = page.extract_tables(table_settings=settings)
                except Exception as exc:
                    warnings.append(f"{pdf_path.name}: page {page.page_number}, extraction table echouee: {exc}")
                    continue

                for table in tables or []:
                    for test_name, spec, results in extract_spec_results_from_table(table):
                        conformity, comment = evaluate_conformity(spec, results)
                        page_rows.append(
                            {
                                "FileName": file_name,
                                "TestName": test_name or "NA",
                                "SPEC": spec,
                                "RESULTS": results,
                                "Conformite": conformity,
                                "Commentaire": comment,
                            }
                        )
                if page_rows:
                    inspection_rows.extend(page_rows)
                    break

    if not inspection_rows:
        inspection_rows.append(
            {
                "FileName": file_name,
                "TestName": "NA",
                "SPEC": "NA",
                "RESULTS": "NA",
                "Conformite": "A VERIFIER",
                "Commentaire": "Aucun tableau SPEC/RESULTS trouve",
            }
        )
        warnings.append(f"{pdf_path.name}: aucun tableau SPEC/RESULTS trouve")

    return PdfReport(
        cover_rows=cover_rows,
        standard_rows=standard_rows,
        inspection_rows=inspection_rows,
        warnings=warnings,
    )


def find_pdfs(input_path: Path) -> list[Path]:
    if input_path.is_file() and input_path.suffix.lower() == ".pdf":
        return [input_path]
    return sorted(input_path.glob("*.pdf"))


def write_csv(rows: list[dict[str, str]], output_path: Path, columns: list[str]) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def autosize_columns(ws) -> None:
    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            max_length = max(max_length, len(str(cell.value or "")))
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 55)


def style_sheet(ws) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    ok_fill = PatternFill("solid", fgColor="C6EFCE")
    warning_fill = PatternFill("solid", fgColor="FFEB9C")
    bad_fill = PatternFill("solid", fgColor="FFC7CE")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            value = str(cell.value or "").upper()
            if value in {"OK", "CONFORME"}:
                cell.fill = ok_fill
            elif value in {"A VERIFIER"}:
                cell.fill = warning_fill
            elif value in {"DIFFERENT", "NON CONFORME"}:
                cell.fill = bad_fill

    ws.freeze_panes = "A2"
    autosize_columns(ws)


def append_sheet(workbook, title: str, rows: list[dict[str, str]], columns: list[str]) -> None:
    ws = workbook.create_sheet(title=title)
    ws.append(columns)
    for row in rows:
        ws.append([row.get(column, "NA") for column in columns])
    style_sheet(ws)


def write_excel_report(
    cover_rows: list[dict[str, str]],
    standard_rows: list[dict[str, str]],
    inspection_rows: list[dict[str, str]],
    output_path: Path,
) -> None:
    from openpyxl import Workbook

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    append_sheet(workbook, "Page de garde", cover_rows, COVER_COLUMNS)
    append_sheet(workbook, "Normes", standard_rows, STANDARD_COLUMNS)
    append_sheet(workbook, "Inspection", inspection_rows, INSPECTION_COLUMNS)
    workbook.save(output_path)


def write_csv_report(
    cover_rows: list[dict[str, str]],
    standard_rows: list[dict[str, str]],
    inspection_rows: list[dict[str, str]],
    output_path: Path,
) -> list[Path]:
    stem = output_path.with_suffix("")
    cover_path = stem.with_name(stem.name + "_page_de_garde.csv")
    standards_path = stem.with_name(stem.name + "_normes.csv")
    inspection_path = stem.with_name(stem.name + "_inspection.csv")

    write_csv(cover_rows, cover_path, COVER_COLUMNS)
    write_csv(standard_rows, standards_path, STANDARD_COLUMNS)
    write_csv(inspection_rows, inspection_path, INSPECTION_COLUMNS)

    if output_path.suffix.lower() == ".csv":
        write_csv(inspection_rows, output_path, INSPECTION_COLUMNS)
        return [output_path, cover_path, standards_path, inspection_path]
    return [cover_path, standards_path, inspection_path]


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Extrait et verifie les certificats qualite PCB depuis des PDF."
    )
    parser.add_argument(
        "input",
        nargs="?",
        default=".",
        help="Dossier contenant les PDF ou chemin d'un PDF unique. Defaut: dossier courant.",
    )
    parser.add_argument(
        "--output",
        default="rapport_pcb.xlsx",
        help="Chemin du rapport .xlsx ou .csv. Defaut: rapport_pcb.xlsx",
    )
    return parser


def generate_report(input_path: Path, output_path: Path) -> tuple[int, dict[str, object]]:
    pdfs = find_pdfs(input_path)
    if not pdfs:
        return 1, {"error": f"Aucun PDF trouve dans: {input_path}"}

    cover_rows: list[dict[str, str]] = []
    standard_rows: list[dict[str, str]] = []
    inspection_rows: list[dict[str, str]] = []
    all_warnings: list[str] = []

    for pdf_path in pdfs:
        print(f"Traitement: {pdf_path.name}")
        report = extract_pdf_report(pdf_path)
        cover_rows.extend(report.cover_rows)
        standard_rows.extend(report.standard_rows)
        inspection_rows.extend(report.inspection_rows)
        all_warnings.extend(report.warnings)

    if output_path.suffix.lower() == ".xlsx":
        write_excel_report(cover_rows, standard_rows, inspection_rows, output_path)
        generated_paths = [output_path]
    else:
        generated_paths = write_csv_report(cover_rows, standard_rows, inspection_rows, output_path)

    summary: dict[str, object] = {
        "pdf_count": len(pdfs),
        "cover_count": len(cover_rows),
        "standard_count": len(standard_rows),
        "inspection_count": len(inspection_rows),
        "warnings": all_warnings,
        "generated_paths": generated_paths,
    }
    return 0, summary


def print_summary(summary: dict[str, object], output_path: Path) -> None:
    generated_paths = summary.get("generated_paths", [])
    if output_path.suffix.lower() == ".xlsx":
        print(f"Rapport Excel genere: {output_path}")
    else:
        print("Rapports CSV generes:")
        for path in generated_paths:
            print(f"- {path}")

    print(f"PDF traites: {summary.get('pdf_count', 0)}")
    print(f"Lignes page de garde: {summary.get('cover_count', 0)}")
    print(f"Normes exportees: {summary.get('standard_count', 0)}")
    print(f"Lignes inspection: {summary.get('inspection_count', 0)}")

    all_warnings = summary.get("warnings", [])
    if all_warnings:
        print("\nAlertes / verifications:")
        for warning in all_warnings:
            print(f"- {warning}")


def main() -> int:
    args = build_arg_parser().parse_args()
    input_path = Path(args.input).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve()

    try:
        status, summary = generate_report(input_path, output_path)
    except ModuleNotFoundError as exc:
        missing = exc.name or "module requis"
        print(f"Module manquant: {missing}")
        print("Installe les dependances avec: python -m pip install pdfplumber openpyxl")
        return 1

    if status:
        print(summary.get("error", "Erreur inconnue"))
        return status

    print_summary(summary, output_path)
    return 0


def launch_gui() -> None:
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk

    root = tk.Tk()
    root.title("Controle qualite PCB")
    root.geometry("720x390")
    root.minsize(680, 360)

    input_var = tk.StringVar()
    output_var = tk.StringVar(value=str(Path.cwd() / "rapport_pcb.xlsx"))
    status_var = tk.StringVar(value="Selectionnez un PDF ou un dossier contenant des PDF.")

    def choose_pdf() -> None:
        path = filedialog.askopenfilename(
            title="Choisir un PDF",
            filetypes=[("Fichiers PDF", "*.pdf"), ("Tous les fichiers", "*.*")],
        )
        if path:
            input_var.set(path)
            output_var.set(str(Path(path).with_name("rapport_pcb.xlsx")))

    def choose_folder() -> None:
        path = filedialog.askdirectory(title="Choisir un dossier de PDF")
        if path:
            input_var.set(path)
            output_var.set(str(Path(path) / "rapport_pcb.xlsx"))

    def choose_output() -> None:
        path = filedialog.asksaveasfilename(
            title="Enregistrer le rapport",
            defaultextension=".xlsx",
            filetypes=[("Rapport Excel", "*.xlsx"), ("CSV", "*.csv")],
            initialfile=Path(output_var.get()).name,
        )
        if path:
            output_var.set(path)

    def set_busy(is_busy: bool) -> None:
        state = "disabled" if is_busy else "normal"
        for button in action_buttons:
            button.configure(state=state)
        progress.configure(mode="indeterminate" if is_busy else "determinate")
        if is_busy:
            progress.start(12)
        else:
            progress.stop()
            progress["value"] = 0

    def run_report() -> None:
        input_text = input_var.get().strip()
        output_text = output_var.get().strip()
        if not input_text:
            messagebox.showwarning("Entrée manquante", "Choisissez un PDF ou un dossier.")
            return
        if not output_text:
            messagebox.showwarning("Sortie manquante", "Choisissez le fichier Excel à générer.")
            return

        input_path = Path(input_text).expanduser().resolve()
        output_path = Path(output_text).expanduser().resolve()
        status_var.set("Traitement en cours...")
        set_busy(True)

        def worker() -> None:
            try:
                status, summary = generate_report(input_path, output_path)
                if status:
                    raise RuntimeError(str(summary.get("error", "Erreur inconnue")))
                warnings = summary.get("warnings", [])
                message = (
                    f"Rapport généré:\n{output_path}\n\n"
                    f"PDF traités: {summary.get('pdf_count', 0)}\n"
                    f"Lignes inspection: {summary.get('inspection_count', 0)}"
                )
                if warnings:
                    message += f"\n\nAlertes: {len(warnings)} ligne(s) à vérifier."

                def done() -> None:
                    set_busy(False)
                    status_var.set("Rapport généré. Ouverture d'Excel...")
                    messagebox.showinfo("Terminé", message)
                    try:
                        os.startfile(output_path)
                    except OSError:
                        status_var.set(f"Rapport généré: {output_path}")

                root.after(0, done)
            except Exception as exc:
                def failed() -> None:
                    set_busy(False)
                    status_var.set("Erreur pendant le traitement.")
                    messagebox.showerror("Erreur", str(exc))

                root.after(0, failed)

        threading.Thread(target=worker, daemon=True).start()

    container = ttk.Frame(root, padding=18)
    container.pack(fill="both", expand=True)

    title = ttk.Label(container, text="Controle qualite PCB", font=("Segoe UI", 17, "bold"))
    title.pack(anchor="w")

    subtitle = ttk.Label(
        container,
        text="Choisissez un PDF ou un dossier. Le rapport Excel sera généré puis ouvert automatiquement.",
        wraplength=650,
    )
    subtitle.pack(anchor="w", pady=(4, 18))

    input_frame = ttk.LabelFrame(container, text="Entrée", padding=12)
    input_frame.pack(fill="x", pady=(0, 12))
    ttk.Entry(input_frame, textvariable=input_var).pack(side="left", fill="x", expand=True, padx=(0, 8))
    pdf_button = ttk.Button(input_frame, text="Choisir PDF", command=choose_pdf)
    folder_button = ttk.Button(input_frame, text="Choisir dossier", command=choose_folder)
    pdf_button.pack(side="left", padx=(0, 6))
    folder_button.pack(side="left")

    output_frame = ttk.LabelFrame(container, text="Rapport", padding=12)
    output_frame.pack(fill="x", pady=(0, 12))
    ttk.Entry(output_frame, textvariable=output_var).pack(side="left", fill="x", expand=True, padx=(0, 8))
    output_button = ttk.Button(output_frame, text="Changer", command=choose_output)
    output_button.pack(side="left")

    run_button = ttk.Button(container, text="Generer et ouvrir le rapport", command=run_report)
    run_button.pack(anchor="e", pady=(4, 12))

    progress = ttk.Progressbar(container)
    progress.pack(fill="x", pady=(0, 8))
    ttk.Label(container, textvariable=status_var).pack(anchor="w")

    action_buttons = [pdf_button, folder_button, output_button, run_button]
    root.mainloop()


if __name__ == "__main__":
    if len(sys.argv) == 1:
        launch_gui()
    else:
        raise SystemExit(main())
